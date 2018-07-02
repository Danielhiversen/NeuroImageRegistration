# -*- coding: utf-8 -*-
"""
Created on Tue Apr 19 15:19:49 2016

@author: dahoiv
"""
# pylint: disable= line-too-long
from __future__ import print_function
import glob
import os
import shutil
import sqlite3
import pyexcel_xlsx
from openpyxl import load_workbook
import nibabel as nib

import image_registration
import util

# DATA_PATH_LISA = MAIN_FOLDER + "Segmenteringer_Lisa/"
# PID_LISA = MAIN_FOLDER + "Koblingsliste__Lisa.xlsx"
# DATA_PATH_LISA_QOL = MAIN_FOLDER + "Segmenteringer_Lisa/Med_QoL/"
# DATA_PATH_ANNE_LISE = MAIN_FOLDER + "Segmenteringer_AnneLine/"
# PID_ANNE_LISE = MAIN_FOLDER + "Koblingsliste__Anne_Line.xlsx"
# DATA_PATH_LGG = MAIN_FOLDER + "Data_HansKristian_LGG/LGG/NIFTI/"

MAIN_FOLDER = "/media/leb/data/"
CONVERTER_PATH = "/home/leb/dev/BRAINSTools/build/bin/ConvertBetweenFileFormats"

def create_db(path):
    """Make the database"""
    conn = sqlite3.connect(path)
    cursor = conn.cursor()

    cursor.execute('''CREATE TABLE "Images" (
    `id`    INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE,
    `pid`    INTEGER,
    `modality`    TEXT,
    `diag_pre_post`    TEXT,
    `filepath`    TEXT,
    `transform`    TEXT,
    `fixed_image`    INTEGER,
    `filepath_reg`    TEXT,
    `comments`    TEXT,
    FOREIGN KEY(`pid`) REFERENCES `Patient`(`pid`))''')
    cursor.execute('''CREATE TABLE "Labels" (
    `id`    INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE,
    `image_id`    INTEGER NOT NULL,
    `description`    TEXT,
    `filepath`    TEXT,
    `filepath_reg`    TEXT,
    `comments`    TEXT,
    FOREIGN KEY(`image_id`) REFERENCES `Images`(`id`))''')
    cursor.execute('''CREATE TABLE "Patient" (
    `pid`    INTEGER NOT NULL UNIQUE,
    `glioma_grade`    INTEGER,
    `comments`    TEXT,
    PRIMARY KEY(pid))''')
    cursor.execute('''CREATE TABLE "QualityOfLife" (
    `id`    INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT UNIQUE,
    `pid`    INTEGER NOT NULL,
    'Index_value'     REAL,
    'Global_index'    INTEGER,
    'Mobility'    INTEGER,
    'Selfcare'    INTEGER,
    'Activity'    INTEGER,
    'Pain'    INTEGER,
    'Anxiety'    INTEGER,
    FOREIGN KEY(`pid`) REFERENCES `Patient`(`pid`))''')

    conn.commit()
    cursor.close()

    conn.close()


def get_convert_table(path):
    """Open xls file and read new pid"""
    xls_data = pyexcel_xlsx.get_data(path)
    convert_table = {}
    data = xls_data['Ark1']
    for row in data:
        if not row:
            continue
        pid = str(row[0])
        case_id = str(row[1])
        convert_table[case_id] = pid
    return convert_table


# pylint: disable= too-many-arguments, too-many-locals
def convert_and_save_dataset(pid, cursor, image_type, volume_labels, volume, glioma_grade):
    """convert_and_save_dataset"""
    util.mkdir_p(util.DATA_FOLDER + str(pid))
    img_out_folder = util.DATA_FOLDER + str(pid) + "/volumes_labels/"
    util.mkdir_p(img_out_folder)

    cursor.execute('''SELECT pid from Patient where pid = ?''', (pid,))
    exist = cursor.fetchone()
    if exist is None:
        cursor.execute('''INSERT INTO Patient(pid, glioma_grade) VALUES(?, ?)''', (pid, glioma_grade))

    cursor.execute('''INSERT INTO Images(pid, modality, diag_pre_post) VALUES(?,?,?)''',
                   (pid, 'MR', image_type))
    img_id = cursor.lastrowid

    _, file_extension = os.path.splitext(volume)
    volume_temp = "volume" + file_extension
    shutil.copy(volume, volume_temp)

    volume_out = img_out_folder + str(pid) + "_" + str(img_id) + "_MR_T1_" + image_type + ".nii.gz"
    print("--->", volume_out)
    os.system(CONVERTER_PATH + " '" + volume_temp + "' " + volume_out)

    volume_out_db = volume_out.replace(util.DATA_FOLDER, "")
    cursor.execute('''UPDATE Images SET filepath = ?, filepath_reg = ? WHERE id = ?''', (volume_out_db, None, img_id))
    os.remove(volume_temp)

    for volume_label in volume_labels:
        _, file_extension = os.path.splitext(volume_label)
        volume_label_temp = "volume_label" + file_extension
        shutil.copy(volume_label, volume_label_temp)

        cursor.execute('''INSERT INTO Labels(image_id, description) VALUES(?,?)''',
                       (img_id, 'all'))
        label_id = cursor.lastrowid

        volume_label_out = img_out_folder + str(pid) + "_" + str(img_id) + "_MR_T1_" + image_type\
            + "_label_all.nii.gz"
        os.system(CONVERTER_PATH + " '" + volume_label_temp + "' " + volume_label_out)
        volume_label_out_db = volume_label_out.replace(util.DATA_FOLDER, "")
        cursor.execute('''UPDATE Labels SET filepath = ?, filepath_reg = ? WHERE id = ?''',
                       (volume_label_out_db, None, label_id))
        os.remove(volume_label_temp)


def qol_to_db(data_type):
    # pylint: disable= too-many-branches
    """Convert qol data to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    data = pyexcel_xlsx.get_data('/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/Indexverdier_atlas_250117.xlsx')['Ark2']

    k = 0
    for row in data:
        k = k + 1
        if not row:
            continue
        if data_type == "extra" and k < 53:
            continue
        if data_type == "siste_runde" and k < 84:
            continue
        elif k < 4:
            continue

        if data_type == "gbm":
            idx1 = 1
            idx2 = 0
            idx3 = 7
        elif data_type in ["lgg", "extra", "siste_runde"]:
            idx1 = 12
            idx2 = 11
            idx3 = 18
        if len(row) < idx3 - 1:
            continue
        print(row)

        if row[idx1] is None:
            gl_idx = None
        elif row[idx1] > 0.85:
            gl_idx = 1
        elif row[idx1] > 0.50:
            gl_idx = 2
        else:
            gl_idx = 3

        val = [None]*7
        idx = 0
        for _val in row[idx2:idx3]:
            val[idx] = _val
            idx += 1
        pid = val[0]
        cursor.execute('''SELECT id from QualityOfLife where pid = ?''', (pid,))
        _id = cursor.fetchone()
        if _id:
            print("-----------", row)
            continue
        cursor.execute('''INSERT INTO QualityOfLife(pid, Index_value, Global_index, Mobility, Selfcare, Activity, Pain, Anxiety) VALUES(?,?,?,?,?,?,?,?)''',
                       (pid, val[1], gl_idx, val[2], val[3], val[4], val[5], val[6]))
        conn.commit()

    cursor.close()
    conn.close()


def qol_change_to_db():
    # pylint: disable= too-many-branches
    """Convert qol data to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    try:
        conn.execute("alter table QualityOfLife add column 'Delta_mobility' 'INTEGER'")
        conn.execute("alter table QualityOfLife add column 'Delta_selfcare' 'INTEGER'")
        conn.execute("alter table QualityOfLife add column 'Delta_activity' 'INTEGER'")
        conn.execute("alter table QualityOfLife add column 'Delta_pain' 'INTEGER'")
        conn.execute("alter table QualityOfLife add column 'Delta_anixety' 'INTEGER'")
        conn.execute("alter table QualityOfLife add column 'Resection' 'INTEGER'")
        conn.execute("alter table QualityOfLife add column 'Delta_qol' 'REAL'")
        conn.execute("alter table QualityOfLife add column 'Delta_kps' 'INTEGER'")
    except sqlite3.OperationalError:
        pass
    sheet = load_workbook('/home/dahoiv/disk/data/Segmentations/Endring_QoL_KPS_1904.xlsx', data_only=True)['Ark1']
    for row in range(2, 223):
        cell_name = "{}{}".format("A", row)
        color = sheet[cell_name].fill.start_color.index
        value = sheet[cell_name].value
        try:
            pid = int(value)
        except (ValueError, TypeError):
            continue
        print(pid, color)
        d_qol = sheet["{}{}".format("B", row)].value
        d_mobility = sheet["{}{}".format("C", row)].value
        d_selfcare = sheet["{}{}".format("D", row)].value
        d_activity = sheet["{}{}".format("E", row)].value
        d_pain = sheet["{}{}".format("F", row)].value
        d_anxity = sheet["{}{}".format("G", row)].value
        resection = 1 if color != 'FFFFFF00' else 0

        cursor.execute('''UPDATE QualityOfLife SET Delta_qol = ? WHERE pid = ?''', (d_qol, pid))
        cursor.execute('''UPDATE QualityOfLife SET Delta_mobility = ? WHERE pid = ?''', (d_mobility, pid))
        cursor.execute('''UPDATE QualityOfLife SET Delta_selfcare = ? WHERE pid = ?''', (d_selfcare, pid))
        cursor.execute('''UPDATE QualityOfLife SET Delta_activity = ? WHERE pid = ?''', (d_activity, pid))
        cursor.execute('''UPDATE QualityOfLife SET Delta_pain = ? WHERE pid = ?''', (d_pain, pid))
        cursor.execute('''UPDATE QualityOfLife SET Delta_anixety = ? WHERE pid = ?''', (d_anxity, pid))
        cursor.execute('''UPDATE QualityOfLife SET Resection = ? WHERE pid = ?''', (resection, pid))
        conn.commit()

        print(d_mobility, d_selfcare, d_activity, d_pain, d_anxity, resection)

    conn.commit()
    cursor.close()
    conn.close()


def karnofsky_to_db():
    """Convert qol data to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    data = pyexcel_xlsx.get_data('/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/Indexverdier_atlas_250117.xlsx')['Ark3']
    try:
        conn.execute("alter table QualityOfLife add column 'karnofsky' 'INTEGER'")
    except sqlite3.OperationalError:
        pass

    k = 0
    for row in data:
        k = k + 1
        if not row:
            continue
        if k < 3:
            continue
        print(row)
        try:
            float(row[0])
        except ValueError:
            continue

        cursor.execute('''UPDATE QualityOfLife SET karnofsky = ? WHERE pid = ?''',
                       (row[1], row[0]))
        conn.commit()

    cursor.close()
    conn.close()


def convert_data(path, glioma_grade, update=False, case_ids=range(2000)):
    """Convert gbm data """
    # pylint: disable= too-many-locals, too-many-branches, too-many-statements
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    log = ""
    
    for case_id in case_ids:
        data_path = path + str(case_id) + "/"
        if not os.path.exists(data_path):
            continue
        
        pid = str(case_id)
        print("\n ====================\n Adding patient " + pid)
        if update:
            for _file in glob.glob(util.DATA_FOLDER + str(pid) + "/volumes_labels/*"):
                _file = _file.replace(util.DATA_FOLDER, "")
                cursor.execute("DELETE FROM Images WHERE filepath=?", (_file,))
                cursor.execute("DELETE FROM Labels WHERE filepath=?", (_file,))
            try:
                shutil.rmtree(util.DATA_FOLDER + str(pid))
            except OSError:
                pass

        file_type_nrrd = True
        volume_label = glob.glob(data_path + '/*label.nrrd')
        if not volume_label:
            volume_label = glob.glob(data_path + '/*label_1.nrrd')
        if not volume_label:
            volume_label = glob.glob(data_path + '/Segmentation/*label.nrrd')
        if len(volume_label) > 1:
            log = log + "\n Warning!! More than one file with label found "
            for volume_label_i in volume_label:
                log = log + volume_label_i
            continue
        if not volume_label:
            file_type_nrrd = False
            volume = data_path + "k" + pid + "-T1_" + "preop" + ".nii"
            if not os.path.exists(volume):
                log = log + "\n Warning!! No volumes found" + data_path + volume

            volume_label = data_path + "k" + pid + "_" + "preop" + "_" + "hele-label" + ".nii"
            if not os.path.exists(volume_label):
                volume_label = glob.glob(data_path + '/*label.nii')[0]
            if not os.path.exists(volume_label):
                log = log + "\n Warning!! No label found " + data_path + volume_label

        if file_type_nrrd:
            volume_label = volume_label[0]
            volume = volume_label.replace("-label", "")
            if not os.path.exists(volume):
                volume = glob.glob(data_path + '*.nrrd')
                volume.remove(volume_label)
                if not volume:
                    volume = glob.glob(data_path + '*.nii')
                if len(volume) > 1:
                    log = log + "\n Warning!! More than one file with volume found " + volume
                if not volume:
                    log = log + "\n Warning!! No volume found " + data_path
                volume = volume[0]

        convert_and_save_dataset(pid, cursor, "pre", [volume_label], volume, glioma_grade)
        conn.commit()

    with open("Log.txt", "w") as text_file:
        text_file.write(log)
    cursor.close()
    conn.close()

    
def convert_lgg_data(path):
    """convert_lgg_data"""
    convert_table = get_convert_table('/home/dahoiv/disk/data/Segmentations/NY_PID_LGG segmentert.xlsx')

    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    # pylint: disable= no-member
    ids = range(70)
    ids.extend(['X1', 'X3', 'X5'])
    print(ids)
    for case_id_org in ids:
        if isinstance(case_id_org, int):
            case_id = '%02d' % case_id_org
        else:
            case_id = str(case_id_org)
        volume = path + case_id + '_post.nii'
        image_type = 'post'
        if not os.path.exists(volume):
            volume = path + case_id + '_pre.nii'
            image_type = 'pre'
        print(volume)
        if not os.path.exists(volume):
            continue

        pid = convert_table[str(case_id_org)]
        print(volume, image_type, case_id, pid)

        if image_type == 'post':
            volume_label = path + case_id + '_post-label.nii'
        elif image_type == 'pre':
            volume_label = path + case_id + '_pre-label.nii'
        if not os.path.exists(volume):
            continue

        convert_and_save_dataset(pid, cursor, image_type, [volume_label], volume, 2)
        conn.commit()

    cursor.close()
    conn.close()


def vacuum_db():
    """ Clean up database"""
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.execute('''VACUUM; ''')
    cursor.close()
    conn.close()


def update_segmentations(path):
    """Update existing patients with new segmentations from Even"""
    # pylint: disable= too-many-locals, too-many-branches, too-many-statements
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    log = ""

    included_cases = path + "Included cases - final.xlsx"
    case_list = load_workbook(included_cases,data_only=True)['Included cases - final']
    for case in range(2, 213):

        cell_name = "{}{}".format("A", case)
        pid = str(case_list[cell_name].value)

        cell_name = "{}{}".format("B", case)
        new_segmentation = str(case_list[cell_name].value)

        data_path = path + "Segmenteringer/" + pid + "/"

        if new_segmentation == "1" and os.path.exists(util.DATA_FOLDER + pid + "/" ):# and pid == "711":
            print("Converting image " + pid)
            log = log + "\n Converting image " + pid
            
            label_file_endings = ('*label.nrrd',
                                  '*label_1.nrrd',
                                  'Segmentation/*label.nrrd',
                                  'Segmentering/*label.nrrd',
                                  '*hele-label.nii')       

            for file_ending in label_file_endings:
                volume_label = glob.glob(data_path + file_ending)    
                if volume_label:
                    break
            if len(volume_label) > 1:
                log = log + "\n Warning!! More than one file with label found "
                for volume_label_i in volume_label:
                    log = log + volume_label_i
                continue
            volume_label = volume_label[0]
            if not os.path.exists(volume_label):
                log = log + "\n Warning!! No label found " + data_path + volume_label
            
            cursor.execute("SELECT id FROM Images WHERE pid = ?", (pid,))
            image_id = cursor.fetchone()
            image_id = str(image_id[0])

            cursor.execute("SELECT transform FROM Images WHERE id = ?", (image_id,))
            transforms_temp = cursor.fetchone()
            if transforms_temp is None:
                log = log + "\n Warning!! No transform found for image " + image_id
            else:
                transforms_temp = str(transforms_temp[0])
            transforms = []
            for _transform in transforms_temp.split(","):
                transforms.append(util.DATA_FOLDER + _transform.strip())
             
            cursor.execute("SELECT filepath, filepath_reg FROM Labels WHERE image_id = ?", (image_id,))
                        
            for row in cursor:
                _filepath = row[0]
                _filepath_reg = row[1]
                try:
                    os.remove(util.DATA_FOLDER + _filepath)
                    os.remove(util.DATA_FOLDER + _filepath_reg)
                except OSError:
                    pass
                
                # Converting file to nii.gz if necessary
                os.system(CONVERTER_PATH + " '" + volume_label + "' " + util.DATA_FOLDER + _filepath)

                temp = util.compress_vol( image_registration.move_vol(util.DATA_FOLDER + _filepath, transforms, True) )
                shutil.copy(temp, util.DATA_FOLDER + _filepath_reg)


    with open("Log.txt", "w") as text_file:
        text_file.write(log)
    cursor.close()
    conn.close()    

    
def update_glioma_grade(glioma_grade):
    """Convert qol data to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()
    convert_table = get_convert_table('/home/dahoiv/disk/data/Segmentations/NY_PID_LGG segmentert.xlsx')
    for pid in convert_table.values():
        try:
            pid = int(pid)
        except ValueError:
            continue
        cursor.execute('''UPDATE PATIENT SET glioma_grade = ? WHERE pid = ?''',
                       (glioma_grade, pid))

    cursor.close()
    conn.commit()
    conn.close()


def update_glioma_grade2(path, glioma_grade):
    """Convert qol data to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    for case_id in range(2000):
        data_path = path + str(case_id) + "/"
        if not os.path.exists(data_path):
            continue
        pid = str(case_id)
        print(pid)
        cursor.execute('''UPDATE PATIENT SET glioma_grade = ? WHERE pid = ?''',
                       (glioma_grade, pid))

    cursor.close()
    conn.commit()
    conn.close()


def manual_add_to_db():
    """Convert qol data to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    #    pid = 462
    #    image_type = 'pre'
    #    volume_labels = ['/home/dahoiv/disk/data/Segmentations/AA_til_3D-atlas_271016/462/04 t2_spc_irprep_ns_sag_dark-fl_p2_iso-label.nrrd']
    #    volume = '/home/dahoiv/disk/data/Segmentations/AA_til_3D-atlas_271016/462/04 t2_spc_irprep_ns_sag_dark-fl_p2_iso.nrrd'
    #    glioma_grade = 3
    #
    #    pid = 1061
    #    image_type = 'pre'
    #    volume_labels = ['/home/dahoiv/disk/data/Segmentations/Data_HansKristian_LGG/LGG_070217/1061_pre/1061_preY-label.nrrd']
    #    volume = '/home/dahoiv/disk/data/Segmentations/Data_HansKristian_LGG/LGG_070217/1061_pre/1061_preY.nrrd'
    #    glioma_grade = 2
    #
    #    pid = 16073
    #    image_type = 'pre'
    #    volume_labels = ['/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16073/68 t1_mprage_sag_p2_iso-_ERIK_label.nrrd']
    #    volume = '/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16073/68 t1_mprage_sag_p2_iso.3.12.2.1107.5.2.43.66087.nrrd'
    #    glioma_grade = 34
    #
    #    pid = 16117
    #    image_type = 'pre'
    #    volume_labels = ['/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16117/pid160122 t2_spc_da-fl_sag_p2_iso_1-alserik_label.nrrd']
    #    volume = '/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16117/pid160122 t2_spc_da-fl_sag_p2_iso_1.0.nrrd'
    #    glioma_grade = 34

    # pid = 16087
    # image_type = 'pre'
    # volume_labels = ['/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16087/04 t1_mpr_ns_sag_p2_iso 1mm_iso-erikkorrigert_label.nrrd']
    # volume = '/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16087/04 t1_mpr_ns_sag_p2_iso 1mm_iso.3.12.2.1107.5.2.19.45157.nrrd'
    # glioma_grade = 34

    pid = 16071
    image_type = 'pre'
    volume_labels = ['/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16071/79 t1_mpr_ns_sag_p2_iso 1mm_iso-embkorrigert_label.nrrd']
    volume = '/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/16071/79 t1_mpr_ns_sag_p2_iso 1mm_iso.3.12.2.1107.5.2.19.45157.nrrd'
    glioma_grade = 34

    convert_and_save_dataset(pid, cursor, image_type, volume_labels, volume, glioma_grade)

    cursor.close()
    conn.commit()
    conn.close()


def add_survival_days():
    """add survival_days to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    data = pyexcel_xlsx.get_data('/home/dahoiv/disk/data/Segmentations/Overlevelse_til_3D_atlas.xlsx')['Ark1']
    try:
        conn.execute("alter table Patient add column 'survival_days' 'INTEGER'")
    except sqlite3.OperationalError:
        pass
    try:
        conn.execute("alter table Patient add column 'op_date' 'INTEGER'")
    except sqlite3.OperationalError:
        pass

    k = 0
    for row in data:
        k = k + 1
        if not row:
            continue
        if k < 2:
            continue
        pid = row[0]
        cursor.execute('''SELECT pid from Patient where pid = ?''', (pid,))
        exist = cursor.fetchone()
        if exist is None:
            continue
        try:
            survival_days = row[3]
        except IndexError:
            op_date = None
        try:
            op_date = row[1]
        except IndexError:
            survival_days = None
        print(pid, survival_days, op_date)

        cursor.execute('''UPDATE Patient SET survival_days = ?, op_date = ? WHERE pid = ?''',
                       (survival_days, op_date, pid))
        conn.commit()

    cursor.close()
    conn.close()


def add_survival_age_kps_days():
    """add survival_days to database """
    # pylint: disable= too-many-branches, too-many-statements
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    data = pyexcel_xlsx.get_data('/home/dahoiv/disk/data/Segmentations/slettes.xlsx')['Ark1']
    try:
        conn.execute("alter table Patient add column 'age_at_op_date_in_days' 'INTEGER'")
    except sqlite3.OperationalError:
        pass

    k = 0
    for row in data:
        k = k + 1
        if not row:
            continue
        if k < 2:
            continue
        pid = row[0]
        cursor.execute('''SELECT pid from Patient where pid = ?''', (pid,))
        exist = cursor.fetchone()
        if exist is None:
            continue
        try:
            op_date = row[1]
        except IndexError:
            op_date = None
        try:
            dob = row[2]
        except IndexError:
            dob = None
        try:
            dod = row[3]
        except IndexError:
            dod = None
        if dod is None:
            try:
                dod = row[4]
            except IndexError:
                dod = None

        if None not in [dod, op_date]:
            survival_days = (dod - op_date).days
        else:
            survival_days = None
        if None not in [dob, op_date]:
            age_at_op_date_in_days = (op_date - dob).days
        else:
            age_at_op_date_in_days = None

        try:
            kps = float(row[6])
        except (IndexError, ValueError):
            kps = None
        print(pid, survival_days, age_at_op_date_in_days, kps)

        cursor.execute('''UPDATE Patient SET survival_days = ?, age_at_op_date_in_days = ? WHERE pid = ?''',
                       (survival_days, age_at_op_date_in_days, pid))
        cursor.execute('''UPDATE QualityOfLife SET karnofsky = ? WHERE pid = ?''', (kps, pid))

        conn.commit()

    cursor.close()
    conn.close()


def add_study():
    """add study to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()
    try:
        conn.execute("alter table Patient add column 'study_id' 'TEXT'")
    except sqlite3.OperationalError:
        pass
    sheet = load_workbook('/home/dahoiv/disk/data/Segmentations/siste_runde_hgg/Indexverdier_atlas_250117.xlsx', data_only=True)['Ark2']
    columns = ["A", "L"]
    k = 0
    for column in columns:
        for row in range(3, 223):
            cell_name = "{}{}".format(column, row)
            color = sheet[cell_name].fill.start_color.index
            value = sheet[cell_name].value
            if value and color == '00000000':
                try:
                    pid = int(value)
                except ValueError:
                    continue
                print(pid, color)
                k += 1
                print(k)
                cursor.execute('''UPDATE Patient SET study_id = ? WHERE pid = ?''',
                               ("qol_grade3,4", pid))
                conn.commit()

    cursor.close()
    conn.close()


def add_study_lgg():
    """add study to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()
    try:
        conn.execute("alter table Patient add column 'study_id' 'TEXT'")
    except sqlite3.OperationalError:
        pass
    sheet = load_workbook('/home/dahoiv/disk/data/Segmentations/pas_til_kart_oversikt.xlsx', data_only=True)['Ark1']
    k = 0
    for row in range(3, 223):
        cell_name = "{}{}".format("A", row)
        # color = sheet[cell_name].fill.start_color.index
        value = sheet[cell_name].value
        try:
            pid = int(value)
        except ValueError:
            continue
        except TypeError:
            continue
        k += 1
        print(pid, k)
        cursor.execute('''UPDATE Patient SET study_id = ? WHERE pid = ?''',
                       ("LGG_reseksjonsgrad", pid))
        conn.commit()

    cursor.close()
    conn.close()


def add_study_survival(path):
    """add study to database """
    conn = sqlite3.connect(util.DB_PATH)

    # Remove commas in study IDs
    cursor = conn.cursor()
    cursor2 = conn.cursor()
    cursor.execute("SELECT pid, study_id FROM Patient")
    for row in cursor:
        pid = row[0]
        study_id = row[1]
        if study_id:
            cursor2.execute("UPDATE Patient SET study_id = ? WHERE pid = ?",
                                   (study_id.replace('qol_grade3,4','qol_grade34'), pid))
            conn.commit()
    cursor2.close()
    
    # Add new study IDs
    survival_id = "GBM_survival_time"
    included_cases = path + "Included cases - final.xlsx"
    case_list = load_workbook(included_cases,data_only=True)['Included cases - final']
    for case in range(2, 213):

        cell_name = "{}{}".format("A", case)
        pid = case_list[cell_name].value

        cursor.execute("SELECT study_id FROM Patient WHERE pid = ?", (pid,))
        study_id = cursor.fetchone()
        if study_id and study_id[0]:
            study_id = study_id[0]
            if survival_id not in study_id:
                study_id += ", " + survival_id
        else:
            study_id = survival_id    
        cursor.execute("UPDATE Patient SET study_id = ? WHERE pid = ?",
                                  (study_id, pid))
        conn.commit()
        
    cursor.close()        
    conn.close()    

def add_survival_in_days(path):
    """add survival data to database """
    
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()
    
    included_cases = path + "Location and survival - survival in days.xlsx"
    case_list = load_workbook(included_cases,data_only=True)['Location and survival - surviva']
    for case in range(2, 213):

        cell_name = "{}{}".format("A", case)
        pid = case_list[cell_name].value

        cell_name = "{}{}".format("B", case)
        survival_days = case_list[cell_name].value

        cursor.execute("SELECT survival_days FROM Patient WHERE pid = ?", (pid,))
        survival_days_db = cursor.fetchone()
        if survival_days_db and survival_days_db[0]:
            if survival_days_db[0] != survival_days:
                print("survival_days_db is not equal to survival_days for pid " + str(pid))
        elif survival_days != "#NULL!":
            print("Survival days for pid " + str(pid) + ": " + str(survival_days))
            cursor.execute("UPDATE Patient SET survival_days = ? WHERE pid = ?",
                              (survival_days, pid))
            conn.commit()
        
    cursor.close()
    conn.close()    
    

def add_tumor_volume():
    """add tumor volume to database """
    conn = sqlite3.connect(util.DB_PATH)
    cursor = conn.cursor()

    data = pyexcel_xlsx.get_data('/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/Volum_kart.xlsx')['Ark1']
    try:
        conn.execute("alter table Images add column 'tumor_volume' 'REAL'")
    except sqlite3.OperationalError:
        pass

    k = 0
    for row in data:
        k = k + 1
        if not row:
            continue
        if k < 2:
            continue
        pid = row[0]
        cursor.execute('''SELECT id from Images where pid = ? AND diag_pre_post="pre"''', (pid,))
        _id = cursor.fetchone()
        if _id is None:
            continue
        _id = _id[0]
        try:
            tumor_volume = row[1]
        except IndexError:
            continue

        print(pid, _id, tumor_volume)

        cursor.execute('''UPDATE Images SET tumor_volume = ? WHERE id = ?''',
                       (tumor_volume, _id))
        conn.commit()

    cursor.close()
    conn.close()


if __name__ == "__main__":
#    util.setup_paths()

    temp_path = "reg_labels_temp/"
    util.setup(temp_path)

    add_survival_in_days(MAIN_FOLDER + "Even_survival/")
#    add_study_survival(MAIN_FOLDER + "Even_survival/")
    
#    update_segmentations(MAIN_FOLDER + "Even_survival/")

#    convert_data(MAIN_FOLDER + "Even_survival/Segmenteringer/New_patients/", 4, update=True, case_ids=range(16000, 17000))






#    try:
#        shutil.rmtree(util.DATA_FOLDER)
#    except OSError:
#        pass
#    util.mkdir_p(util.DATA_FOLDER)
#    create_db(util.DB_PATH)
#    convert_data(MAIN_FOLDER + "Segmenteringer_GBM/", 4)
#    qol_to_db("gbm")

#    convert_lgg_data(MAIN_FOLDER + "Data_HansKristian_LGG/LGG/NIFTI/PRE_OP/")
#    convert_lgg_data(MAIN_FOLDER + "Data_HansKristian_LGG/LGG/NIFTI/POST/")
#    qol_to_db("lgg")

#    qol_to_db("extra")
#    convert_data(MAIN_FOLDER + "LGG_til_3D-atlas_271016/", 2, True)
#    convert_data(MAIN_FOLDER + "AA_til_3D-atlas_271016/", 3)

#    convert_data(MAIN_FOLDER + "GBM_til_3D-atlas_revidert_031116/", 4, True)
#    karnofsky_to_db()

#    update_glioma_grade(2)

#    update_glioma_grade2(MAIN_FOLDER + "LGG_til_3D-atlas_271016/", 2)
#    update_glioma_grade2(MAIN_FOLDER + "AA_til_3D-atlas_271016/", 3)
#    update_glioma_grade2(MAIN_FOLDER + "GBM_til_3D-atlas_revidert_031116/", 4)
#    update_glioma_grade2(MAIN_FOLDER + "Segmenteringer_GBM/", 4)

#    manual_add_to_db()

#    add_survival_days()
#    manual_add_to_db()

#    convert_data(MAIN_FOLDER + "siste_runde_hgg/", 34, update=True)
#    convert_data(MAIN_FOLDER + "siste_runde_hgg/", 34, update=False, case_ids=range(2000, 20000))
#    add_study()
#    karnofsky_to_db()
#    convert_data(MAIN_FOLDER + "siste_runde_hgg/", 34, update=True, case_ids=[1424])

#    qol_to_db("siste_runde")

#    qol_change_to_db()

#    add_survival_age_kps_days()


#    add_study_lgg()

#    add_tumor_volume()

    vacuum_db()
