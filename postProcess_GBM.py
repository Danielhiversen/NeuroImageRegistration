# -*- coding: utf-8 -*-
"""
Created on Tue May 24 10:41:50 2016

@author: dahoiv
"""

import os
import nipype.interfaces.slicer as slicer
from openpyxl import Workbook
import datetime
import numpy as np
import nibabel as nib
import sqlite3

import util
import do_img_registration_GBM


BRAINSResample_PATH = '/home/dahoiv/disk/kode/Slicer/Slicer-SuperBuild/Slicer-build/lib/Slicer-4.6/cli-modules/BRAINSResample'


def find_images():
    """ Find images for registration """
    conn = sqlite3.connect(util.DB_PATH)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from Patient where study_id = ?''', ("qol_grade3,4", ))
    ids = []
    for row in cursor:
        cursor3 = conn.execute('''SELECT Resection from QualityOfLife where pid = ?''', (row[0], ))
        resection = cursor3.fetchone()[0]
        cursor3.close()
        if resection in [0, None]:
            continue
        cursor2 = conn.execute('''SELECT id from Images where pid = ?''', (row[0], ))
        for _id in cursor2:
            ids.append(_id[0])
        cursor2.close()

    cursor.close()
    conn.close()
    return ids


def find_images_163():
    """ Find images for registration """
    conn = sqlite3.connect(util.DB_PATH)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from Patient where study_id = ?''', ("qol_grade3,4", ))
    ids = []
    k = 0
    for row in cursor:
        k += 1
        cursor3 = conn.execute('''SELECT Index_value from QualityOfLife where pid = ?''', (row[0], ))
        indx_val = cursor3.fetchone()[0]
        cursor3.close()
        if indx_val in [None]:
            continue

        cursor2 = conn.execute('''SELECT id from Images where pid = ?''', (row[0], ))
        for _id in cursor2:
            ids.append(_id[0])
        cursor2.close()

    cursor.close()
    conn.close()
    return ids


def process_vlsm(folder, n_permutations):
    """ Post process vlsm data """
    print(folder)
    util.setup(folder)
    image_ids = find_images()
    params = ['Index_value', 'karnofsky',  'Mobility', 'Selfcare', 'Activity', 'Pain', 'Anxiety']
    alternative = ['less', 'less', 'greater', 'greater', 'greater', 'greater', 'greater']
    stat_func = [util.brunner_munzel_test, util.mannwhitneyu_test, util.mannwhitneyu_test, util.mannwhitneyu_test,
                 util.mannwhitneyu_test, util.mannwhitneyu_test, util.mannwhitneyu_test]
    for (qol_param, stat_func_i, alternative_i) in zip(params, stat_func, alternative):
        (image_ids_with_qol, qol) = util.get_qol(image_ids, qol_param)
        result = util.post_calculations(image_ids_with_qol)
        for label in result:
            print(label)
            if label == 'img':
                continue
            util.vlsm(result[label], label + '_' + qol_param, stat_func_i, qol, folder,
                      n_permutations=n_permutations, alternative=alternative_i)


def process(folder):
    """ Post process data distribution and baseline"""
    print(folder)
    util.setup(folder)
    params = ['Mobility', 'Selfcare', 'Activity', 'Pain', 'Anxiety', 'karnofsky', 'Index_value']
    image_ids = find_images_163()
    result = util.post_calculations(image_ids)
    print(len(result['all']))
    util.avg_calculation(result['all'], 'all_N=163', None, True, folder, save_sum=True)
    util.avg_calculation(result['img'], 'img_N=163', None, True, folder)

    image_ids = do_img_registration_GBM.find_images()
    result = util.post_calculations(image_ids)
    print(len(result['all']))
    util.avg_calculation(result['all'], 'all_N=170', None, True, folder, save_sum=True)
    util.avg_calculation(result['img'], 'img_N=170', None, True, folder)
    for qol_param in params:
        (image_ids_with_qol, qol) = util.get_qol(image_ids, qol_param)
        if qol_param not in ["karnofsky", "Delta_kps"]:
            qol = [_temp * 100 for _temp in qol]
        default_value = -100
        print(qol_param)
        print(len(qol))
        result = util.post_calculations(image_ids_with_qol)
        for label in result:
            if label == 'img':
                continue
            print(label)
            # util.avg_calculation(result[label], label + '_' + qol_param, qol, True, folder, default_value=default_value)
            util.median_calculation(result[label], label + '_' + qol_param, qol, True, folder, default_value=default_value)
            # util.std_calculation(result[label], label + '_' + qol_param, qol, True, folder)


def process2(folder):
    """ Post process data Delta"""
    print(folder)
    util.setup(folder)
    params = ['Delta_qol', 'Delta_qol2', 'Delta_mobility', 'Delta_selfcare', 'Delta_activity', 'Delta_pain', 'Delta_anixety', 'Delta_kps']
    image_ids = find_images()
    result = util.post_calculations(image_ids)
    print(len(result['all']))
    util.avg_calculation(result['all'], 'all_N=112', None, True, folder, save_sum=True)
    util.avg_calculation(result['img'], 'img_N=112', None, True, folder)
    print("\n\n\n\n\n")

    for qol_param in params:
        if qol_param == "Delta_qol2":
            (image_ids_with_qol, qol) = util.get_qol(image_ids, "Delta_qol")
            qol = [-1 if _temp <= -0.15 else 0 if _temp < 0.15 else 1 for _temp in qol]
        else:
            (image_ids_with_qol, qol) = util.get_qol(image_ids, qol_param)
        qol = [_temp * 100 for _temp in qol]
        default_value = -300
        print(qol_param, len(qol))
        result = util.post_calculations(image_ids_with_qol)
        for label in result:
            if label == 'img':
                continue
            print(label)
            # util.avg_calculation(result[label], label + '_' + qol_param, qol, True, folder, default_value=default_value)
            util.median_calculation(result[label], label + '_' + qol_param, qol, True, folder, default_value=default_value)
            # util.std_calculation(result[label], label + '_' + qol_param, qol, True, folder)


def process3(folder):
    """ Post process data """
    print(folder)
    util.setup(folder)
    params = ['Mobility', 'Selfcare', 'Activity', 'Pain', 'Anxiety', 'karnofsky', 'Index_value']
    image_ids = find_images()
    result = util.post_calculations(image_ids)
    print(len(result['all']))
    # util.avg_calculation(result['all'], 'all_N=112', None, True, folder, save_sum=True)
    # util.avg_calculation(result['img'], 'img_N=112', None, True, folder)

    for qol_param in params:
        if qol_param == "Delta_qol2":
            (image_ids_with_qol, qol) = util.get_qol(image_ids, "Delta_qol")
            qol = [-1 if _temp <= -0.15 else 0 if _temp < 0.15 else 1 for _temp in qol]
        else:
            (image_ids_with_qol, qol) = util.get_qol(image_ids, qol_param)
        if qol_param not in ["karnofsky", "Delta_kps"]:
            qol = [_temp * 100 for _temp in qol]
        default_value = -100
        print(qol_param)
        print(len(qol))
        result = util.post_calculations(image_ids_with_qol)
        for label in result:
            if label == 'img':
                continue
            print(label)
            util.avg_calculation(result[label], label + '_' + qol_param + '_N=112', qol, True, folder, default_value=default_value)
            util.median_calculation(result[label], label + '_' + qol_param + '_N=112', qol, True, folder, default_value=default_value)
            # util.std_calculation(result[label], label + '_' + qol_param, qol, True, folder)


def process4(folder):
    """ Post process data tumor volume"""
    print(folder)
    util.setup(folder)
    default_value = 0
    label = 'all'

    for image_ids in [do_img_registration_GBM.find_images(), find_images(), find_images_163()]:
        result = util.post_calculations(image_ids)
        (image_ids_with_qol, qol) = util.get_tumor_volume(image_ids)
        num = len(result['all'])
        print(num)
        util.median_calculation(result[label], 'tumor_volume_N=' + str(num), qol, True, folder, default_value=default_value)


def process_labels(folder):
    """ Post process data tumor volume"""
    print(folder)
    util.setup(folder)
    conn = sqlite3.connect(util.DB_PATH, timeout=120)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from Patient where study_id = ?''', ("qol_grade3,4", ))

    img = nib.load("/home/dahoiv/disk/data/MolekylareMarkorer/lobes_brain.nii")
    lobes_brain = img.get_data()
    label_defs = util.get_bigger_label_defs()
    label_defs_r_l = util.get_right_left_label_defs()
    res_lobes_brain = {}

    book = Workbook()
    sheet = book.active

    sheet.cell(row=1, column=1).value = 'PID'
    sheet.cell(row=1, column=2).value = 'Lobe'
    sheet.cell(row=1, column=3).value = 'Right/Left'
    # sheet.cell(row=1, column=3).value = 'Center of mass'
    k = 2
    for pid in cursor:
        pid = pid[0]

        _id = conn.execute('''SELECT id from Images where pid = ?''', (pid, )).fetchone()
        if not _id:
            print("---No data for ", pid)
            continue
        _id = _id[0]

        _filepath = conn.execute("SELECT filepath_reg from Labels where image_id = ?",
                                 (_id, )).fetchone()[0]
        if _filepath is None:
            print("No filepath for ", pid)
            continue

        com, com_idx = util.get_center_of_mass(util.DATA_FOLDER + _filepath)

        lobe = label_defs.get(lobes_brain[com_idx[0], com_idx[1], com_idx[2]], 'other')
        res_lobes_brain[pid] = lobe

        sheet.cell(row=k, column=1).value = pid
        sheet.cell(row=k, column=2).value = lobe
        sheet.cell(row=k, column=3).value = 'left' if com_idx[0] < 99 else 'right'
        rl = label_defs_r_l.get(lobes_brain[com_idx[0], com_idx[1], com_idx[2]], 'other')
        if rl != 'unknown' and rl != sheet.cell(row=k, column=3).value:
            print("\n\n\n", pid)
        # sheet.cell(row=k, column=3).value = str(com[0]) + " " + str(com[1]) + " " + str(com[2])
        # sheet.cell(row=k, column=4).value = str(com_idx[0]) + " " + str(com_idx[1]) + " " + str(com_idx[2])
        k += 1

    book.save("brain_lobes.xlsx")

    print(res_lobes_brain, len(res_lobes_brain))


def process_labels2(folder):
    """ Post process data tumor volume"""
    print(folder)
    util.setup(folder)
    conn = sqlite3.connect(util.DB_PATH, timeout=120)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from Patient where study_id = ?''', ("qol_grade3,4", ))

    atlas_path = "/home/dahoiv/disk/Dropbox/Jobb/gbm/Atlas/Hammers/Hammers_mith-n30r95-MaxProbMap-full-MNI152-SPM12.nii.gz"
    resample = slicer.registration.brainsresample.BRAINSResample(command=BRAINSResample_PATH,
                                                                 inputVolume=atlas_path,
                                                                 outputVolume=os.path.abspath(folder +
                                                                                              'Hammers_mith-n30r95-MaxProbMap-full'
                                                                                              '-MNI152-SPM12_resample.nii.gz'),
                                                                 referenceVolume=os.path.abspath(util.TEMPLATE_VOLUME))
    resample.run()

    img = nib.load(folder + 'Hammers_mith-n30r95-MaxProbMap-full-MNI152-SPM12_resample.nii.gz')
    lobes_brain = img.get_data()
    label_defs = util.get_label_defs_hammers_mith()
    res_lobes_brain = {}

    book = Workbook()
    sheet = book.active

    sheet.cell(row=1, column=1).value = 'PID'
    sheet.cell(row=1, column=2).value = 'Lobe'
    # sheet.cell(row=1, column=3).value = 'Center of mass'
    k = 2
    for pid in cursor:
        pid = pid[0]

        _id = conn.execute('''SELECT id from Images where pid = ?''', (pid, )).fetchone()
        if not _id:
            print("---No data for ", pid)
            continue
        _id = _id[0]

        _filepath = conn.execute("SELECT filepath_reg from Labels where image_id = ?",
                                 (_id, )).fetchone()[0]
        if _filepath is None:
            print("No filepath for ", pid)
            continue

        com, com_idx = util.get_center_of_mass(util.DATA_FOLDER + _filepath)

        lobe = label_defs.get(lobes_brain[com_idx[0], com_idx[1], com_idx[2]], 'other')
        res_lobes_brain[pid] = lobe

        sheet.cell(row=k, column=1).value = pid
        sheet.cell(row=k, column=2).value = lobe
        sheet.cell(row=k, column=4).value = str(com_idx[0]) + " " + str(com_idx[1]) + " " + str(com_idx[2])
        k += 1

    book.save("brain_lobes_Hammers_mith_n30r95.xlsx")

    print(res_lobes_brain, len(res_lobes_brain))


def process_tracts(folder):
    """ Post process data tumor volume"""
    util.setup(folder)
    print(folder)
    thres = 0.75
    atlas_paths = ["/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Arcuate/Arcuate_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Arcuate/Arcuate_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Anterior_Commissure/Anterior_Commissure.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cerebellar/Cortico_Ponto_Cerebellum_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cerebellar/Cortico_Ponto_Cerebellum_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cerebellar/Inferior_Cerebellar_Pedunculus_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cerebellar/Inferior_Cerebellar_Pedunculus_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cerebellar/Superior_Cerebelar_Pedunculus_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cerebellar/Superior_Cerebelar_Pedunculus_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cingulum/Cingulum_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Cingulum/Cingulum_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Corpus_Callosum/Corpus_Callosum.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Fornix/Fornix.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Inferior_Network/Inferior_Longitudinal_Fasciculus_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Inferior_Network/Inferior_Longitudinal_Fasciculus_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Inferior_Network/Inferior_Occipito_Frontal_Fasciculus_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Inferior_Network/Inferior_Occipito_Frontal_Fasciculus_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Inferior_Network/Uncinate_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Inferior_Network/Uncinate_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Optic_Radiations/Optic_Radiations_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Optic_Radiations/Optic_Radiations_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Perisylvian/Anterior_Segment_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Perisylvian/Anterior_Segment_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Perisylvian/Long_Segment_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Perisylvian/Long_Segment_Right.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Perisylvian/Posterior_Segment_Left.nii",
                   "/mnt/b7cde2db-ac2d-4cbb-b2b0-a9b110f05d32/data/Segmentations/WM_tracts/Perisylvian/Posterior_Segment_Right.nii",
                   "/home/dahoiv/disk/Dropbox/Jobb/gbm/FINAL_RES_GBM_0919_09_06_2017/WM_tracts/Projections/Internal_Capsule.nii"
                   ]

    for atlas_path in atlas_paths:
        tract = util.get_basename(atlas_path)
        resample = slicer.registration.brainsresample.BRAINSResample(command=BRAINSResample_PATH,
                                                                     inputVolume=atlas_path,
                                                                     outputVolume=os.path.abspath(folder + tract + '.nii.gz'),
                                                                     referenceVolume=os.path.abspath(util.TEMPLATE_VOLUME))
        print(resample.cmdline)
        resample.run()

    conn = sqlite3.connect(util.DB_PATH, timeout=120)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from Patient where study_id = ?''', ("qol_grade3,4", ))

    book = Workbook()
    sheet = book.active

    sheet.cell(row=1, column=1).value = 'PID'

    k = 2
    for pid in cursor:
        pid = pid[0]

        _id = conn.execute('''SELECT id from Images where pid = ?''', (pid, )).fetchone()
        if not _id:
            print("---No data for ", pid)
            continue
        _id = _id[0]

        _filepath = conn.execute("SELECT filepath_reg from Labels where image_id = ?",
                                 (_id, )).fetchone()[0]
        if _filepath is None:
            print("No filepath for ", pid)
            continue

        tumor_data = nib.load(util.DATA_FOLDER + _filepath).get_data()

        sheet.cell(row=k, column=1).value = pid
        m = 1
        for atlas_path in atlas_paths:
            tract = util.get_basename(atlas_path)
            if 'Internal_Capsule' not in tract:
                continue
            m += 1
            sheet.cell(row=1, column=l).value = tract
            atlas_data = nib.load(folder + tract + '.nii.gz').get_data()
            union_data = atlas_data * tumor_data

            sheet.cell(row=k, column=m).value = '1' if np.max(union_data) >= thres else '0'
        k += 1

    book.save("brain_tracts_Internal_Capsule.xlsx")


if __name__ == "__main__":
    folder = "RES_GBM_" + "{:%H%M_%m_%d_%Y}".format(datetime.datetime.now()) + "/"
    # process(folder)
    # process2(folder)
    # # process3(folder)
    # process4(folder)
    # process_labels(folder)
    process_labels2(folder)
    # process_tracts(folder)

    # start_time = datetime.datetime.now()
    # if len(sys.argv) > 1:
    #     n_permutations = int(sys.argv[1])
    # else:
    #     n_permutations = 20
    # # process_vlsm(folder, n_permutations)
    # print("Total runtime")
    # print(datetime.datetime.now() - start_time)
