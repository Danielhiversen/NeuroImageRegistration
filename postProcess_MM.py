# -*- coding: utf-8 -*-
"""
Created on Tue May 24 10:41:50 2016

@author: dahoiv
"""

# import os
from openpyxl import Workbook
import collections
#import nipype.interfaces.slicer as slicer
import nipype.interfaces.semtools.registration.brainsresample as brainsresample
import pickle
import datetime
import util
import sqlite3
import nibabel as nib
import numpy as np
from scipy.spatial import distance
import os


def format_dict(d):
    d = collections.OrderedDict(sorted(d.iteritems()))
    s = ['lobe                   Type1   Type2   Type3 \n']
    for k, v in d.items():
        v = str(v[0]) + "      " + str(v[1]) + "      " + str(v[2])
        tab = 25 - len(k)
        s.append('%s%s %s\n' % (k, ' '*tab,  v))
    return ''.join(s) + '\n\n'


def process(folder, pids_to_exclude=()):
    """ Post process data """
    print(folder)
    util.setup(folder, 'MolekylareMarkorer')
    conn = sqlite3.connect(util.DB_PATH, timeout=120)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from MolekylareMarkorer ORDER BY pid''')
    image_ids = []
    image_ids_1 = []
    image_ids_2 = []
    image_ids_3 = []
    tag_data_1 = []
    tag_data_2 = []
    tag_data_3 = []
    img = nib.load("/media/leb/data/Atlas/lobes_brain.nii")
    lobes_brain = img.get_data()
    label_defs = util.get_label_defs()
    res_right_left_brain = {}
    res_lobes_brain = {}
    patients = '\nPID  MM\n----------------\n'

    for pid in cursor:
        pid = pid[0]
        if pid in pids_to_exclude:
            continue

        _id = conn.execute('''SELECT id from Images where pid = ?''', (pid, )).fetchone()
        if not _id:
            print("---No data for ", pid)
            continue
        _id = _id[0]

        _mm = conn.execute("SELECT Subgroup from MolekylareMarkorer where pid = ?",
                           (pid, )).fetchone()[0]
        if _mm is None:
            print("No mm data for ", pid)
            patients += str(pid) + ': ?\n'
            continue

        _desc = conn.execute("SELECT comments from MolekylareMarkorer where pid = ?",
                             (pid, )).fetchone()[0]
        if _desc is None:
            _desc = ""

        _filepath = conn.execute("SELECT filepath_reg from Labels where image_id = ?",
                                 (_id, )).fetchone()[0]
        if _filepath is None:
            print("No filepath for ", pid)
            continue

        com, com_idx = util.get_center_of_mass(util.DATA_FOLDER + _filepath)
        val = {}
        val['Name'] = str(pid) + "_" + str(_mm)
        val['PositionGlobal'] = str(com[0]) + "," + str(com[1]) + "," + str(com[2])
        val['desc'] = str(_desc)

        lobe = label_defs[lobes_brain[com_idx[0], com_idx[1], com_idx[2]]]
        right_left = 'left' if com_idx[0] < 99 else 'right'
        res_lobes_brain[lobe] = res_lobes_brain.get(lobe, [0, 0, 0])
        res_right_left_brain[right_left] = res_right_left_brain.get(right_left, [0, 0, 0])
        print(right_left, lobe)
        if _mm == 1:
            res_lobes_brain[lobe][0] += 1
            res_right_left_brain[right_left][0] += 1
            image_ids_1.extend([_id])
        elif _mm == 2:
            res_lobes_brain[lobe][1] += 1
            res_right_left_brain[right_left][1] += 1
            image_ids_2.extend([_id])
        elif _mm == 3:
            res_lobes_brain[lobe][2] += 1
            res_right_left_brain[right_left][2] += 1
            image_ids_3.extend([_id])

        image_ids.extend([_id])
        print(pid, _mm)
        patients += str(pid) + ': ' + str(_mm) + '\n'
        if _mm == 1:
            tag_data_1.append(val)
        elif _mm == 2:
            tag_data_2.append(val)
        elif _mm == 3:
            tag_data_3.append(val)

    print(format_dict(res_lobes_brain))
    lobes_brain_file = open(folder + "lobes_brain.txt", 'w')
    lobes_brain_file.write(format_dict(res_lobes_brain))
    lobes_brain_file.close()
    lobes_brain_file = open(folder + "lobes_brain.txt", 'a')
    lobes_brain_file.write(format_dict(res_right_left_brain))
    lobes_brain_file.write(patients)
    lobes_brain_file.close()

    print(len(image_ids))

    result = util.post_calculations(image_ids_1)
    util.avg_calculation(result['all'], 'mm_1', None, True, folder, save_sum=True)

    result = util.post_calculations(image_ids_2)
    util.avg_calculation(result['all'], 'mm_2', None, True, folder, save_sum=True)

    result = util.post_calculations(image_ids_3)
    util.avg_calculation(result['all'], 'mm_3', None, True, folder, save_sum=True)

    result = util.post_calculations(image_ids)
    util.avg_calculation(result['all'], 'mm_1_2_3', None, True, folder, save_sum=True)

    return
    tag_data = {"tag_data_1": tag_data_1, "tag_data_2": tag_data_2, "tag_data_3": tag_data_3}
    pickle.dump(tag_data, open("tag_data.pickle", "wb"))

    cursor.close()
    conn.close()
    util.write_fcsv("mm_1", folder, tag_data_1, '1 0 0', 13)
    util.write_fcsv("mm_2", folder, tag_data_2, '0 1 0', 5)
    util.write_fcsv("mm_3", folder, tag_data_3, '0 0 1', 6)
    result = util.post_calculations(image_ids)
    util.avg_calculation(result['all'], 'all', None, True, folder, save_sum=True)
    util.avg_calculation(result['img'], 'img', None, True, folder)


def process_labels(folder, pids_to_exclude=()):
    """ Post process data tumor volume"""
    print(folder)
    util.setup(folder, 'MolekylareMarkorer')
    conn = sqlite3.connect(util.DB_PATH, timeout=120)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from MolekylareMarkorer ORDER BY pid''')

    atlas_path = util.ATLAS_FOLDER_PATH + 'Hammers/Hammers_mith-n30r95-MaxProbMap-full-MNI152-SPM12.nii.gz'
    atlas_resampled_path = folder + 'Hammers_mith-n30r95-MaxProbMap-full-MNI152-SPM12_resample.nii.gz'
    resample = brainsresample.BRAINSResample(command=util.BRAINSResample_PATH,
                                             inputVolume=atlas_path,
                                             outputVolume=os.path.abspath(atlas_resampled_path),
                                             referenceVolume=os.path.abspath(util.TEMPLATE_VOLUME))
    resample.run()

    img = nib.load(atlas_resampled_path)
    lobes_brain = img.get_data()
    label_defs = util.get_label_defs_hammers_mith()
    res_lobes_brain = {}

    coordinates_svz = util.get_label_coordinates(util.ATLAS_FOLDER_PATH + 'SubventricularZone.nii.gz')
    surface_dg = util.get_surface(util.ATLAS_FOLDER_PATH + 'DentateGyrus.nii.gz')

    book = Workbook()
    sheet = book.active

    sheet.cell(row=1, column=1).value = 'PID'
    sheet.cell(row=1, column=2).value = 'MM'
    sheet.cell(row=1, column=3).value = 'Lobe, center of tumor'
    sheet.cell(row=1, column=4).value = 'Distance from SVZ to center of tumor (mm)'
    sheet.cell(row=1, column=5).value = 'Distance from SVZ to border of tumor (mm)'
    sheet.cell(row=1, column=6).value = 'Distance from DG to center of tumor (mm)'
    sheet.cell(row=1, column=7).value = 'Distance from DG to border of tumor (mm)'
    sheet.cell(row=1, column=8).value = 'Tumor volume (mm^3)'
    i = 8
    label_defs_to_column = {}
    for key in label_defs:
        i += 1
        sheet.cell(row=1, column=i).value = label_defs[key]
        label_defs_to_column[key] = i
    k = 2
    for pid in cursor:
        pid = pid[0]

        if pid in pids_to_exclude:
            continue

        _id = conn.execute('''SELECT id from Images where pid = ?''', (pid, )).fetchone()
        if not _id:
            print("---No data for ", pid)
            continue
        _id = _id[0]

        _filepath = conn.execute("SELECT filepath_reg from Labels where image_id = ?",
                                 (_id, )).fetchone()[0]
        if _filepath is None:
            print("No filepath for ", pid)
            continue

        com, com_idx = util.get_center_of_mass(util.DATA_FOLDER + _filepath)
        surface = util.get_surface(util.DATA_FOLDER + _filepath)

        print(pid, com_idx)

        dist_from_svz_to_com = distance.cdist(coordinates_svz, [com], 'euclidean').min()
        dist_from_svz_to_border = distance.cdist(coordinates_svz, surface['point_cloud'], 'euclidean').min()
        dist_from_dg_to_com = util.get_min_distance(surface_dg, [com])
        dist_from_dg_to_border = util.get_min_distance(surface_dg, surface['point_cloud'])

        lobe = label_defs.get(lobes_brain[com_idx[0], com_idx[1], com_idx[2]], 'other')
        res_lobes_brain[pid] = lobe

        img = nib.load(util.DATA_FOLDER + _filepath)
        tumor_data = img.get_data()
        voxel_size = img.get_zooms()
        voxel_volume = np.prod(voxel_size[0:3])
        n_voxels = (tumor_data > 0).sum()
        tumor_volume = n_voxels*voxel_volume

        union_data = lobes_brain * tumor_data
        union_data = union_data.flatten()
        lobe_overlap = ''
        for column in range(1, 1+max(label_defs_to_column.values())):
            sheet.cell(row=k, column=column).value = 0
        for _lobe in np.unique(union_data):
            column = label_defs_to_column.get(_lobe)
            if column is None:
                continue
            sheet.cell(row=k, column=column).value = 1
            lobe_overlap += label_defs.get(_lobe, '') + ', '

        _mm = conn.execute("SELECT Subgroup from MolekylareMarkorer where pid = ?",
                           (pid, )).fetchone()[0]

        sheet.cell(row=k, column=1).value = pid
        sheet.cell(row=k, column=2).value = str(_mm)
        sheet.cell(row=k, column=3).value = lobe
        sheet.cell(row=k, column=4).value = round(dist_from_svz_to_com,2)
        sheet.cell(row=k, column=5).value = round(dist_from_svz_to_border,2)
        sheet.cell(row=k, column=6).value = round(dist_from_dg_to_com,2)
        sheet.cell(row=k, column=7).value = round(dist_from_dg_to_border,2)
        sheet.cell(row=k, column=8).value = round(tumor_volume,1)

        k += 1

    book.save(folder + "brain_lobes_Hammers_mith_n30r95.xlsx")

    print(res_lobes_brain, len(res_lobes_brain))


def validate(folder):
    """ Post process data tumor volume"""
    print(folder)
    util.setup(folder, 'MolekylareMarkorer')
    conn = sqlite3.connect(util.DB_PATH, timeout=120)
    conn.text_factory = str
    cursor = conn.execute('''SELECT pid from MolekylareMarkorer ORDER BY pid''')

    brain_mask = nib.load(util.TEMPLATE_MASK).get_data()

    max_val = 0
    max_pid = -1
    for pid in cursor:
        pid = pid[0]

        _id = conn.execute('''SELECT id from Images where pid = ?''', (pid, )).fetchone()
        if not _id:
            print("---No data for ", pid)
            continue
        _id = _id[0]

        _filepath = conn.execute("SELECT filepath_reg from Labels where image_id = ?",
                                 (_id, )).fetchone()[0]

        tumor_data = nib.load(util.DATA_FOLDER + _filepath).get_data()
        union_data = (1-brain_mask) * tumor_data
        print(pid, np.sum(union_data[:]))
        if np.sum(union_data[:]) > max_val:
            max_val = np.sum(union_data[:])
            max_pid = pid
    print("---------- ", max_pid, max_val)


if __name__ == "__main__":
    folder = "RES_MM_" + "{:%Y%m%d_%H%M}".format(datetime.datetime.now()) + "/"

    pids_to_exclude = (122,148)
    process_labels(folder, pids_to_exclude)
    process(folder, pids_to_exclude)
    # validate(folder)
