# -*- coding: utf-8 -*-
"""
Created on Tue May 24 10:41:50 2016

@author: dahoiv
"""

from openpyxl import Workbook
import nipype.interfaces.semtools.registration.brainsresample as brainsresample
import datetime
import util
import sqlite3
import nibabel as nib
import numpy as np
import os


def process(folder, censor_date):
    """ Post process data """
    print(folder)
    util.setup(folder)

    image_ids, survival_days = util.get_image_id_and_survival_days(study_id="GBM_survival_time",censor_date_str=censor_date)
    result = util.post_calculations(image_ids)
    print('Total: ' + str(len(result['all'])) + ' patients')
    util.avg_calculation(result['all'], 'tumor', None, True, folder, save_sum=True)
    util.mortality_rate_calculation(result['all'], '_all_year', survival_days, True, folder, default_value=-1, max_value=150, per_year=True)
    util.avg_calculation(result['img'], 'volume', None, True, folder)

    image_ids, survival_days = util.get_image_id_and_survival_days(study_id="GBM_survival_time",censor_date_str=censor_date,resection=True)
    result = util.post_calculations(image_ids)
    print('Resected: ' + str(len(result['all'])) + ' patients')
    util.avg_calculation(result['all'], 'tumor_resected', None, True, folder, save_sum=True)
    util.avg_calculation(result['img'], 'volume_resected', None, True, folder)
    util.mortality_rate_calculation(result['all'], '_resected_year', survival_days, True, folder, default_value=-1, max_value=150, per_year=True)

    survival_groups = [
        [0, 182],
        [183, 730],
        [731, float('Inf')]
        ]

    for group in survival_groups:
        image_ids, survival_days = util.get_image_id_and_survival_days(study_id="GBM_survival_time",censor_date_str=censor_date,survival_group=group)
        result = util.post_calculations(image_ids)
        print('Group ' + str(group) + ': ' + str(len(result['all'])) + ' patients')
        label = 'tumor_' + str(group[0]) + '-' + str(group[1])
        util.avg_calculation(result['all'], label, None, True, folder, save_sum=True)



def process_labels(folder, pids_to_exclude=None):
    """ Create Excel-document with overview of which brain areas are affected by the tumors"""
    print(folder)
    util.setup(folder)
    conn = sqlite3.connect(util.DB_PATH, timeout=120)
    conn.text_factory = str
    pids, image_ids = util.get_pids_and_image_ids(study_id="GBM_survival_time",exclude_pid=pids_to_exclude)


    atlas_path = util.ATLAS_FOLDER_PATH + "Hammers/Hammers_mith-n30r95-MaxProbMap-full-MNI152-SPM12.nii.gz"
    atlas_resampled_path = folder + 'Hammers_mith-n30r95-MaxProbMap-full-MNI152-SPM12_resample.nii.gz'
    resample = brainsresample.BRAINSResample(command=util.BRAINSResample_PATH,
                                             inputVolume=atlas_path,
                                             outputVolume=os.path.abspath(atlas_resampled_path),
                                             referenceVolume=os.path.abspath(util.TEMPLATE_VOLUME))
    resample.run()

    img = nib.load(atlas_resampled_path)
    lobes_brain = img.get_data()
    label_defs = util.get_label_defs_hammers_mith()
    res_lobes_brain = {}

    ventricle_label = 49
    com_ventricle, com_idx_ventricle =  util.get_center_of_mass(folder + 'Hammers_mith-n30r95-MaxProbMap-full-MNI152-SPM12_resample.nii.gz',ventricle_label)

    book = Workbook()
    sheet = book.active

    sheet.cell(row=1, column=1).value = 'PID'
    sheet.cell(row=1, column=2).value = 'Lobe, center of tumor'
    sheet.cell(row=1, column=3).value = 'Distance from center of 3rd ventricle to center of tumor (mm)'
    sheet.cell(row=1, column=4).value = 'Distance from center of 3rd ventricle to border of tumor (mm)'
    i = 4
    label_defs_to_column = {}
    for key in label_defs:
        i += 1
        sheet.cell(row=1, column=i).value = label_defs[key]
        label_defs_to_column[key] = i
    # sheet.cell(row=1, column=3).value = 'Center of mass'
    k = 2
    for (pid, _id) in zip(pids, image_ids):

        _filepath = conn.execute("SELECT filepath_reg from Labels where image_id = ?",
                                 (_id, )).fetchone()[0]
        if _filepath is None:
            print("No filepath for ", pid)
            continue

        com, com_idx = util.get_center_of_mass(util.DATA_FOLDER + _filepath)
        surface = util.get_surface(util.DATA_FOLDER + _filepath)

        print(pid, com_idx)

        dist_from_ventricle_to_com = np.linalg.norm(np.array(com)-np.array(com_ventricle))
        dist_from_ventricle_to_border = util.get_min_distance(surface, [com_ventricle])

        lobe = label_defs.get(lobes_brain[com_idx[0], com_idx[1], com_idx[2]], 'other')
        res_lobes_brain[pid] = lobe

        tumor_data = nib.load(util.DATA_FOLDER + _filepath).get_data()
        union_data = lobes_brain * tumor_data
        union_data = union_data.flatten()
        lobe_overlap = ''
        for column in range(1, 1+max(label_defs_to_column.values())):
            sheet.cell(row=k, column=column).value = 0
        for _lobe in np.unique(union_data):
            column = label_defs_to_column.get(_lobe)
            if column is None:
                continue
            sheet.cell(row=k, column=column).value = 1
            lobe_overlap += label_defs.get(_lobe, '') + ', '

        sheet.cell(row=k, column=1).value = pid
        sheet.cell(row=k, column=2).value = lobe
        sheet.cell(row=k, column=3).value = round(dist_from_ventricle_to_com,2)
        sheet.cell(row=k, column=4).value = round(dist_from_ventricle_to_border,2)

        k += 1

    book.save(folder + "brain_lobes_Hammers_mith_n30r95.xlsx")

    print(res_lobes_brain, len(res_lobes_brain))

if __name__ == "__main__":
    folder = "RES_survival_time_" + "{:%Y%m%d_%H%M}".format(datetime.datetime.now()) + "/"
    censor_date = "2018-12-31"
    process_labels(folder)
    process(folder,censor_date)
